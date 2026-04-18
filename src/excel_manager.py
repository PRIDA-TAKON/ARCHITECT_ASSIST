import pandas as pd
import openpyxl
import os
from typing import Dict, Any, List

class ExcelManager:
    def __init__(self, file_path: str):
        self.file_path = file_path
        if not os.path.exists(file_path):
            # Create a blank excel if it doesn't exist
            df = pd.DataFrame()
            df.to_excel(file_path, index=False)
        self.wb = openpyxl.load_file(file_path)

    def read_sheet(self, sheet_name: str = None) -> pd.DataFrame:
        """Reads a specific sheet into a pandas DataFrame."""
        return pd.read_excel(self.file_path, sheet_name=sheet_name)

    def write_data(self, data: Dict[str, List[Any]], sheet_name: str = "Sheet1"):
        """Writes data to a specific sheet."""
        df = pd.DataFrame(data)
        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    def update_cell(self, sheet_name: str, row: int, col: int, value: Any):
        """Updates a specific cell's value."""
        wb = openpyxl.load_workbook(self.file_path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
        sheet.cell(row=row, column=col).value = value
        wb.save(self.file_path)

    def find_and_replace_column_value(self, sheet_name: str, column_name: str, search_val: str, new_val: Any):
        """Finds a value in a column and updates it in the same row."""
        df = self.read_sheet(sheet_name)
        if column_name in df.columns:
            df.loc[df[column_name] == search_val, column_name] = new_val
            self.write_data(df.to_dict('list'), sheet_name=sheet_name)
        else:
            raise ValueError(f"Column '{column_name}' not found in sheet '{sheet_name}'")

if __name__ == "__main__":
    print("Excel Manager module initialized.")
